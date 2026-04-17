#!/usr/bin/env python3
"""H1B LCA FY2026 Q1 streaming parser.

Reads ~570 MB sheet1.xml row-by-row via openpyxl's read_only mode.
Filters: CASE_STATUS == 'Certified'
Aggregates: SOC_CODE × WORKSITE_STATE × WORKSITE_CITY → list of annualised wages.
Output: one JSON file with MSA-level p25/p50/p75 by SOC prefix.

Wages converted to annual:
  Hour  → ×2080
  Week  → ×52
  Bi-Weekly → ×26
  Month → ×12
  Year  → ×1
"""
import json
import os
import statistics
import sys
from collections import defaultdict
from openpyxl import load_workbook

SRC = "data/clean-pipeline/sources/h1b/LCA_Disclosure_Data_FY2026_Q1.xlsx"
OUT = "data/clean-pipeline/output/h1b-salary-anchors.json"

WAGE_MULT = {
    "Hour": 2080, "Hourly": 2080,
    "Week": 52, "Weekly": 52,
    "Bi-Weekly": 26,
    "Month": 12, "Monthly": 12,
    "Year": 1, "Yearly": 1, "Annual": 1,
}

# 25-ish WhichCity profession SOC prefixes (BLS SOC 2018 major groups relevant)
# We store per full SOC code; mapping to profession done downstream.
# US 20 top MSAs → collapse WORKSITE_CITY+STATE to MSA label
MSA_MAP = {
    # California
    ("SAN FRANCISCO", "CA"): "San Francisco-Oakland-Berkeley",
    ("OAKLAND", "CA"): "San Francisco-Oakland-Berkeley",
    ("BERKELEY", "CA"): "San Francisco-Oakland-Berkeley",
    ("SAN MATEO", "CA"): "San Francisco-Oakland-Berkeley",
    ("SAN JOSE", "CA"): "San Jose-Sunnyvale-Santa Clara",
    ("SUNNYVALE", "CA"): "San Jose-Sunnyvale-Santa Clara",
    ("SANTA CLARA", "CA"): "San Jose-Sunnyvale-Santa Clara",
    ("MOUNTAIN VIEW", "CA"): "San Jose-Sunnyvale-Santa Clara",
    ("PALO ALTO", "CA"): "San Jose-Sunnyvale-Santa Clara",
    ("CUPERTINO", "CA"): "San Jose-Sunnyvale-Santa Clara",
    ("LOS ANGELES", "CA"): "Los Angeles-Long Beach-Anaheim",
    ("LONG BEACH", "CA"): "Los Angeles-Long Beach-Anaheim",
    ("ANAHEIM", "CA"): "Los Angeles-Long Beach-Anaheim",
    ("IRVINE", "CA"): "Los Angeles-Long Beach-Anaheim",
    ("SAN DIEGO", "CA"): "San Diego-Chula Vista",
    # New York
    ("NEW YORK", "NY"): "New York-Newark-Jersey City",
    ("BROOKLYN", "NY"): "New York-Newark-Jersey City",
    ("JERSEY CITY", "NJ"): "New York-Newark-Jersey City",
    ("NEWARK", "NJ"): "New York-Newark-Jersey City",
    # Texas
    ("AUSTIN", "TX"): "Austin-Round Rock",
    ("HOUSTON", "TX"): "Houston-The Woodlands-Sugar Land",
    ("DALLAS", "TX"): "Dallas-Fort Worth-Arlington",
    ("FORT WORTH", "TX"): "Dallas-Fort Worth-Arlington",
    ("PLANO", "TX"): "Dallas-Fort Worth-Arlington",
    ("IRVING", "TX"): "Dallas-Fort Worth-Arlington",
    ("SAN ANTONIO", "TX"): "San Antonio-New Braunfels",
    # NW
    ("SEATTLE", "WA"): "Seattle-Tacoma-Bellevue",
    ("BELLEVUE", "WA"): "Seattle-Tacoma-Bellevue",
    ("REDMOND", "WA"): "Seattle-Tacoma-Bellevue",
    ("TACOMA", "WA"): "Seattle-Tacoma-Bellevue",
    # MA
    ("BOSTON", "MA"): "Boston-Cambridge-Newton",
    ("CAMBRIDGE", "MA"): "Boston-Cambridge-Newton",
    # DC
    ("WASHINGTON", "DC"): "Washington-Arlington-Alexandria",
    ("ARLINGTON", "VA"): "Washington-Arlington-Alexandria",
    ("ALEXANDRIA", "VA"): "Washington-Arlington-Alexandria",
    ("MCLEAN", "VA"): "Washington-Arlington-Alexandria",
    ("RESTON", "VA"): "Washington-Arlington-Alexandria",
    # IL
    ("CHICAGO", "IL"): "Chicago-Naperville-Elgin",
    ("NAPERVILLE", "IL"): "Chicago-Naperville-Elgin",
    # GA
    ("ATLANTA", "GA"): "Atlanta-Sandy Springs-Alpharetta",
    # FL
    ("MIAMI", "FL"): "Miami-Fort Lauderdale-West Palm Beach",
    ("FORT LAUDERDALE", "FL"): "Miami-Fort Lauderdale-West Palm Beach",
    # AZ
    ("PHOENIX", "AZ"): "Phoenix-Mesa-Chandler",
    ("MESA", "AZ"): "Phoenix-Mesa-Chandler",
    ("CHANDLER", "AZ"): "Phoenix-Mesa-Chandler",
    # CO
    ("DENVER", "CO"): "Denver-Aurora-Lakewood",
    ("BOULDER", "CO"): "Denver-Aurora-Lakewood",
    # MN
    ("MINNEAPOLIS", "MN"): "Minneapolis-St. Paul-Bloomington",
    # NC
    ("RALEIGH", "NC"): "Raleigh-Cary",
    ("CHARLOTTE", "NC"): "Charlotte-Concord-Gastonia",
    # OR
    ("PORTLAND", "OR"): "Portland-Vancouver-Hillsboro",
    ("HILLSBORO", "OR"): "Portland-Vancouver-Hillsboro",
    # PA
    ("PHILADELPHIA", "PA"): "Philadelphia-Camden-Wilmington",
    ("PITTSBURGH", "PA"): "Pittsburgh",
}


def main():
    print(f"Opening {SRC} (streaming, read_only)...", flush=True)
    wb = load_workbook(SRC, read_only=True, data_only=True)
    ws = wb.active
    print(f"Sheet: {ws.title}, max_row={ws.max_row}, max_col={ws.max_column}", flush=True)

    rows_iter = ws.iter_rows(values_only=True)
    header = [str(c) if c is not None else "" for c in next(rows_iter)]
    print("Header cols:", len(header), flush=True)
    for i, h in enumerate(header):
        if any(k in h.upper() for k in ("STATUS", "SOC", "WAGE", "WORKSITE", "PW_WAGE_LEVEL", "EMPLOYER_NAME")):
            print(f"  [{i}] {h}")
    sys.stdout.flush()

    # index lookups
    idx = {h: i for i, h in enumerate(header)}
    need = [
        "CASE_STATUS", "SOC_CODE", "SOC_TITLE",
        "WORKSITE_CITY", "WORKSITE_STATE",
        "WAGE_RATE_OF_PAY_FROM", "WAGE_UNIT_OF_PAY",
        "PW_WAGE_LEVEL",
    ]
    missing = [k for k in need if k not in idx]
    if missing:
        print("MISSING cols:", missing, file=sys.stderr)
        print("Available:", [h for h in header if h], file=sys.stderr)
        sys.exit(2)

    # aggregate: (msa, soc4) -> list[annual_wage]
    # soc4 = first 5 chars of SOC_CODE e.g. "15-1252" → we keep full code for precision
    buckets = defaultdict(list)
    total, certified, with_msa, bad_wage = 0, 0, 0, 0
    for row in rows_iter:
        total += 1
        if total % 50000 == 0:
            print(f"  scanned {total}  certified={certified}  msa_rows={with_msa}", flush=True)
        status = row[idx["CASE_STATUS"]]
        if status != "Certified":
            continue
        certified += 1
        city = str(row[idx["WORKSITE_CITY"]] or "").strip().upper()
        state = str(row[idx["WORKSITE_STATE"]] or "").strip().upper()
        msa = MSA_MAP.get((city, state))
        if not msa:
            continue
        with_msa += 1
        soc = str(row[idx["SOC_CODE"]] or "").strip()
        title = str(row[idx["SOC_TITLE"]] or "").strip()
        wage = row[idx["WAGE_RATE_OF_PAY_FROM"]]
        unit = str(row[idx["WAGE_UNIT_OF_PAY"]] or "").strip()
        mult = WAGE_MULT.get(unit)
        if wage is None or mult is None:
            bad_wage += 1
            continue
        try:
            w = float(wage) * mult
        except (TypeError, ValueError):
            bad_wage += 1
            continue
        if w < 20000 or w > 2_000_000:  # sanity
            continue
        level = row[idx["PW_WAGE_LEVEL"]]
        buckets[(msa, soc, title)].append((w, level))

    print(f"\nTotal rows:      {total}", flush=True)
    print(f"Certified:       {certified}", flush=True)
    print(f"With MSA match:  {with_msa}", flush=True)
    print(f"Bad wage:        {bad_wage}", flush=True)
    print(f"Buckets:         {len(buckets)}", flush=True)

    # summarise → {msa: {soc: {title, n, p25, p50, p75, levels:{1:n,2:n,...}}}}
    out = {}
    for (msa, soc, title), arr in buckets.items():
        if len(arr) < 5:
            continue
        wages = [a[0] for a in arr]
        levels = defaultdict(int)
        for _, lv in arr:
            levels[str(lv) if lv is not None else "?"] += 1
        wages.sort()
        def q(p):
            k = (len(wages) - 1) * p
            f = int(k)
            c = min(f + 1, len(wages) - 1)
            return wages[f] + (wages[c] - wages[f]) * (k - f)
        row = {
            "title": title,
            "n": len(wages),
            "p25": round(q(0.25)),
            "p50": round(q(0.50)),
            "p75": round(q(0.75)),
            "mean": round(statistics.fmean(wages)),
            "levels": dict(levels),
        }
        out.setdefault(msa, {})[soc] = row

    meta = {
        "source": "US DOL ETA H1B LCA Disclosure Data (Public Domain)",
        "period": "FY2026 Q1 (Oct-Dec 2025)",
        "url": "https://www.dol.gov/agencies/eta/foreign-labor/performance",
        "filter": "CASE_STATUS=Certified, annualized wage 20k–2M, n>=5 per MSA×SOC",
        "units": "USD/year",
        "msaCount": len(out),
        "socBuckets": sum(len(v) for v in out.values()),
    }
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w") as f:
        json.dump({"meta": meta, "byMsa": out}, f, indent=2, ensure_ascii=False)
    print(f"\n✓ wrote {OUT}", flush=True)
    print(f"  {len(out)} MSAs, {sum(len(v) for v in out.values())} MSA×SOC buckets")

if __name__ == "__main__":
    main()
