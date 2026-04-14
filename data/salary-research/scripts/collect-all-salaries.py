#!/usr/bin/env python3
"""
collect-all-salaries.py — WhichCity 全球薪资数据采集脚本（主脚本）

运行环境：任意可联网环境（如果 BLS 被封请通过 --skip-bls 跳过）
        需要 Python 3.9+, openpyxl

用法:
  python3 data/salary-research/scripts/collect-all-salaries.py              # 全量采集
  python3 data/salary-research/scripts/collect-all-salaries.py --skip-bls   # 跳过 BLS（被封时）
  python3 data/salary-research/scripts/collect-all-salaries.py --bls-only   # 仅 BLS

采集阶段:
  Phase 1: BLS OEWS — 美国 21 城市 × 24 职业 (download.bls.gov 文本格式)
  Phase 2: Eurostat SES — 欧盟 ~30 国 × ISCO-08 2位数职业
  Phase 3: World Bank GNI — 用于薪资换算的 PPP/Atlas 基准线
  Phase 4: 已有 ILO 数据的再处理

所有原始数据保存到 data/salary-research/raw/ 目录。
"""

import json
import os
import sys
import time
import urllib.request
import urllib.error
import ssl
import csv
import io
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent.parent.parent
RAW_DIR = ROOT / "data" / "salary-research" / "raw"
MAPPINGS_DIR = ROOT / "data" / "salary-research" / "mappings"
RAW_DIR.mkdir(parents=True, exist_ok=True)

SKIP_BLS = "--skip-bls" in sys.argv
BLS_ONLY = "--bls-only" in sys.argv

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

# ═══════════════════════════════════════════════════════════════════
#  Profession → SOC code mapping
# ═══════════════════════════════════════════════════════════════════
PROFESSION_SOC = {
    "软件工程师":     ["15-1252"],
    "医生/医学博士":  ["29-1228", "29-1211", "29-1216"],
    "财务分析师":     ["13-2051"],
    "市场经理":       ["11-2021"],
    "平面设计师":     ["27-1024"],
    "数据科学家":     ["15-2051"],
    "销售经理":       ["11-2022"],
    "人力资源经理":   ["11-3121"],
    "教师":           ["25-2031"],
    "护士":           ["29-1141"],
    "律师":           ["23-1011"],
    "建筑师":         ["17-1011"],
    "厨师":           ["35-1011"],
    "记者":           ["27-3023"],
    "机械工程师":     ["17-2141"],
    "药剂师":         ["29-1051"],
    "会计师":         ["13-2011"],
    "产品经理":       ["11-2021"],   # proxy: Marketing Manager
    "UI/UX设计师":    ["15-1255"],
    "大学教授":       ["25-1099"],
    "牙医":           ["29-1021"],
    "公交司机":       ["53-3052"],
    "电工":           ["47-2111"],
    "政府/NGO行政":   ["13-1111"],
}

# City ID → BLS MSA code
CITY_MSA = {
    1: ("35620", "New York"),
    11: ("31080", "Los Angeles"),
    12: ("41860", "San Francisco"),
    13: ("16980", "Chicago"),
    34: ("33100", "Miami"),
    35: ("47900", "Washington DC"),
    36: ("14460", "Boston"),
    37: ("42660", "Seattle"),
    38: ("19740", "Denver"),
    39: ("12420", "Austin"),
    73: ("41980", "San Juan PR"),
    95: ("12060", "Atlanta"),
    96: ("38060", "Phoenix"),
    97: ("38900", "Portland"),
    98: ("41740", "San Diego"),
    99: ("29820", "Las Vegas"),
    100: ("45300", "Tampa"),
    125: ("26420", "Houston"),
    126: ("37980", "Philadelphia"),
    133: ("41940", "San Jose"),
    134: ("31080", "Irvine (LA MSA)"),
}

# Reverse lookup: SOC code (without dash) → profession name(s)
SOC_TO_PROF = {}
for prof, codes in PROFESSION_SOC.items():
    for code in codes:
        clean = code.replace("-", "")
        if clean not in SOC_TO_PROF:
            SOC_TO_PROF[clean] = []
        SOC_TO_PROF[clean].append(prof)


def fetch(url, desc="", timeout=30):
    """Fetch URL with error handling and retries."""
    for attempt in range(3):
        try:
            req = urllib.request.Request(url, headers=HEADERS)
            ctx = ssl.create_default_context()
            with urllib.request.urlopen(req, timeout=timeout, context=ctx) as resp:
                return resp.read()
        except (urllib.error.HTTPError, urllib.error.URLError, TimeoutError) as e:
            print(f"  Attempt {attempt+1}/3 failed: {e}")
            if attempt < 2:
                time.sleep(2 ** attempt)
    return None


# ═══════════════════════════════════════════════════════════════════
#  Phase 1: BLS OEWS (US metro areas)
# ═══════════════════════════════════════════════════════════════════

def phase1_bls():
    """
    Download BLS OEWS text-format data from download.bls.gov.
    This server uses a different domain than www.bls.gov and may work.
    If not, fall back to the zip download.
    """
    print("\n" + "=" * 60)
    print("PHASE 1: BLS OEWS (US metro area salary data)")
    print("=" * 60)

    output_file = RAW_DIR / "bls-oews-extracted.json"

    # Strategy A: Try the text-format time series data
    # These are tab-separated files at download.bls.gov
    # File: oe.data.0.Current has format: series_id\tyear\tperiod\tvalue\tfootnotes
    # Series ID format for OEWS:
    #   OEUM  = prefix (metro)
    #   NNNNNNN = 7-digit area code
    #   IIIIII = 6-digit industry (000000 = cross-industry)
    #   OOOOOO = 6-digit SOC code (no dash)
    #   DD = data type (04 = annual mean, 13 = annual median)

    # First try: download the series catalog to understand format
    print("\n  Trying download.bls.gov text format...")
    
    # Download the area codes file
    area_data = fetch("https://download.bls.gov/pub/time.series/oe/oe.area", "area codes")
    if area_data:
        area_text = area_data.decode("utf-8")
        print(f"  ✓ Got area codes ({len(area_text)} bytes)")
        
        # Parse area codes — find our MSA codes
        msa_area_map = {}  # area_code → msa_code
        for line in area_text.strip().split("\n")[1:]:  # skip header
            parts = line.strip().split("\t")
            if len(parts) >= 3:
                area_type = parts[0].strip()
                area_code = parts[1].strip()
                area_name = parts[2].strip() if len(parts) > 2 else ""
                # Metro areas have area_type = 'M'
                for city_id, (msa, city_name) in CITY_MSA.items():
                    if msa in area_code or msa in area_name:
                        msa_area_map[area_code] = (msa, city_name, city_id)
        
        print(f"  Found {len(msa_area_map)} matching area codes")
        
        # Download the data type codes
        datatype_data = fetch("https://download.bls.gov/pub/time.series/oe/oe.datatype", "data types")
        if datatype_data:
            print(f"  ✓ Got data types")
            for line in datatype_data.decode().strip().split("\n"):
                if "median" in line.lower() or "annual" in line.lower():
                    print(f"    {line.strip()}")
        
        # Now try to download actual wage data
        # The main data file can be huge. Let's try the series file first.
        print("\n  Downloading main data file (may be large)...")
        data_raw = fetch(
            "https://download.bls.gov/pub/time.series/oe/oe.data.0.Current",
            "main data",
            timeout=120
        )
        
        if data_raw:
            data_text = data_raw.decode("utf-8", errors="replace")
            lines = data_text.strip().split("\n")
            print(f"  ✓ Got {len(lines):,} data lines ({len(data_raw):,} bytes)")
            
            # Parse: Extract annual median wages for our SOC codes in our MSAs
            # Series ID structure: OEUM + area(7) + industry(6) + occupation(6) + datatype(2)
            # data type 13 = annual median
            
            # Build set of target SOC codes (without dash)
            target_socs = set()
            for codes in PROFESSION_SOC.values():
                for c in codes:
                    target_socs.add(c.replace("-", ""))
            
            # Build set of target MSA codes
            target_msas = set(msa for msa, _ in CITY_MSA.values())
            
            results = []
            matched = 0
            
            for line in lines[1:]:  # skip header
                parts = line.strip().split("\t")
                if len(parts) < 4:
                    continue
                
                series_id = parts[0].strip()
                year = parts[1].strip()
                value = parts[3].strip()
                
                # Only want OEUM prefix (metro data)
                if not series_id.startswith("OEUM"):
                    continue
                
                # Parse series_id components
                # OEUM + season(1) + area(7) + industry(6) + occ(6) + datatype(2)
                # Total length: 4 + 1 + 7 + 6 + 6 + 2 = 26
                if len(series_id) < 26:
                    continue
                
                prefix = series_id[:4]  # OEUM
                seasonal = series_id[4]  # S or U
                area = series_id[5:12]   # 7-digit area
                industry = series_id[12:18]  # 6-digit industry
                occ = series_id[18:24]   # 6-digit occupation
                datatype = series_id[24:26]  # 2-digit data type
                
                # We want: cross-industry (000000), annual median (13), latest year
                if industry != "000000" or datatype != "13":
                    continue
                
                # Check if this MSA and SOC are in our targets
                # Area code in BLS is padded differently. Let's match flexibly
                area_clean = area.lstrip("0")
                msa_match = None
                for msa in target_msas:
                    if area.endswith(msa) or area_clean == msa:
                        msa_match = msa
                        break
                
                if not msa_match:
                    continue
                
                # Check SOC code
                if occ not in target_socs:
                    continue
                
                # Parse value
                try:
                    wage = float(value.replace(",", ""))
                except ValueError:
                    continue
                
                if wage <= 0:
                    continue
                
                # Find city and profession
                city_matches = [(cid, name) for cid, (m, name) in CITY_MSA.items() if m == msa_match]
                prof_matches = SOC_TO_PROF.get(occ, [])
                
                for city_id, city_name in city_matches:
                    for prof in prof_matches:
                        results.append({
                            "cityId": city_id,
                            "cityName": city_name,
                            "profession": prof,
                            "soc": occ[:2] + "-" + occ[2:],
                            "msa": msa_match,
                            "annualMedian": wage,
                            "year": year,
                            "seriesId": series_id,
                        })
                        matched += 1
            
            print(f"\n  Extracted {matched} data points")
            
            if results:
                output = {
                    "source": "US Bureau of Labor Statistics — OEWS",
                    "url": "https://download.bls.gov/pub/time.series/oe/",
                    "collectedAt": time.strftime("%Y-%m-%dT%H:%M:%S"),
                    "method": "download.bls.gov text-format time series",
                    "dataType": "Pre-tax annual median wage (USD)",
                    "totalResults": len(results),
                    "data": results,
                }
                with open(output_file, "w") as f:
                    json.dump(output, f, indent=2, ensure_ascii=False)
                print(f"  ✓ Saved → {output_file}")
                
                # Coverage summary
                cities_found = set(r["cityId"] for r in results)
                profs_found = set(r["profession"] for r in results)
                print(f"\n  Coverage: {len(cities_found)} cities × {len(profs_found)} professions")
                
                for city_id, (msa, city_name) in sorted(CITY_MSA.items()):
                    city_results = [r for r in results if r["cityId"] == city_id]
                    profs = set(r["profession"] for r in city_results)
                    missing = set(PROFESSION_SOC.keys()) - profs
                    status = "✓" if len(missing) == 0 else f"missing: {', '.join(list(missing)[:3])}..."
                    print(f"    {city_name:20s}: {len(profs)}/24 {status}")
                
                return True
            else:
                print("  ✗ No matching data found in BLS text file")
        else:
            print("  ✗ Could not download BLS text data (IP may be blocked)")
    else:
        print("  ✗ Could not reach download.bls.gov")
    
    # Strategy B: Try direct zip download
    print("\n  Trying BLS zip download (oesm24ma.zip)...")
    zip_url = "https://www.bls.gov/oes/special-requests/oesm24ma.zip"
    zip_path = RAW_DIR / "oesm24ma.zip"
    
    zip_data = fetch(zip_url, "BLS OEWS zip", timeout=120)
    if zip_data and len(zip_data) > 100000:
        with open(zip_path, "wb") as f:
            f.write(zip_data)
        print(f"  ✓ Downloaded {len(zip_data):,} bytes → {zip_path}")
        print("  → Run Phase 1B to parse this zip file")
        return True
    else:
        print("  ✗ BLS zip download also blocked")
        print("\n  ╔════════════════════════════════════════════════════╗")
        print("  ║ BLS is blocking this IP.                          ║")
        print("  ║ Please download manually via browser:              ║")
        print("  ║                                                    ║")
        print("  ║ https://www.bls.gov/oes/special-requests/oesm24ma.zip")
        print("  ║                                                    ║")
        print("  ║ Save to: data/salary-research/raw/oesm24ma.zip    ║")
        print("  ║ Then re-run this script.                           ║")
        print("  ╚════════════════════════════════════════════════════╝")
        return False


def phase1b_parse_bls_zip():
    """Parse BLS OEWS zip file if it was downloaded manually."""
    zip_path = RAW_DIR / "oesm24ma.zip"
    if not zip_path.exists():
        return False
    
    print("\n  Parsing BLS OEWS zip file...")
    import zipfile
    
    with zipfile.ZipFile(zip_path) as zf:
        names = zf.namelist()
        print(f"  Files in zip: {names}")
        
        # Find the Excel file
        xlsx_files = [n for n in names if n.endswith('.xlsx')]
        if not xlsx_files:
            print("  ✗ No .xlsx file found in zip")
            return False
        
        xlsx_name = xlsx_files[0]
        print(f"  Parsing {xlsx_name}...")
        
        import openpyxl
        with zf.open(xlsx_name) as xlsx_file:
            wb = openpyxl.load_workbook(xlsx_file, read_only=True, data_only=True)
            ws = wb.active
            
            # Find header row
            headers = {}
            for row in ws.iter_rows(min_row=1, max_row=5, values_only=False):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        headers[cell.value.strip().upper()] = cell.column - 1
                if "AREA" in " ".join(str(h) for h in headers.keys()):
                    break
            
            print(f"  Headers: {list(headers.keys())[:10]}...")
            
            # Expected columns: AREA, AREA_TITLE, OCC_CODE, OCC_TITLE, 
            #                   A_MEDIAN (annual median), TOT_EMP, etc.
            col_area = headers.get("AREA", headers.get("AREA_CODE", None))
            col_occ = headers.get("OCC_CODE", None)
            col_median = headers.get("A_MEDIAN", headers.get("ANNUAL_MEDIAN", None))
            col_mean = headers.get("A_MEAN", headers.get("ANNUAL_MEAN", None))
            col_emp = headers.get("TOT_EMP", None)
            
            if col_area is None or col_occ is None:
                print(f"  ✗ Could not find required columns. Available: {list(headers.keys())}")
                return False
            
            # Parse data rows
            target_msas = {msa for msa, _ in CITY_MSA.values()}
            target_socs = set()
            for codes in PROFESSION_SOC.values():
                for c in codes:
                    target_socs.add(c)
            
            results = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                area = str(row[col_area]).strip() if row[col_area] else ""
                occ = str(row[col_occ]).strip() if row[col_occ] else ""
                
                if area not in target_msas or occ not in target_socs:
                    continue
                
                median_val = row[col_median] if col_median is not None else None
                mean_val = row[col_mean] if col_mean is not None else None
                
                try:
                    wage = float(str(median_val).replace(",", "")) if median_val and str(median_val).replace(",", "").replace(".", "").isdigit() else None
                except ValueError:
                    wage = None
                
                if wage and wage > 0:
                    city_matches = [(cid, name) for cid, (m, name) in CITY_MSA.items() if m == area]
                    prof_matches = SOC_TO_PROF.get(occ.replace("-", ""), [])
                    
                    for city_id, city_name in city_matches:
                        for prof in prof_matches:
                            results.append({
                                "cityId": city_id,
                                "cityName": city_name,
                                "profession": prof,
                                "soc": occ,
                                "msa": area,
                                "annualMedian": wage,
                            })
            
            wb.close()
            
            if results:
                output_file = RAW_DIR / "bls-oews-extracted.json"
                output = {
                    "source": "US Bureau of Labor Statistics — OEWS (May 2024)",
                    "url": "https://www.bls.gov/oes/special-requests/oesm24ma.zip",
                    "collectedAt": time.strftime("%Y-%m-%dT%H:%M:%S"),
                    "method": "XLSX flat file parse",
                    "dataType": "Pre-tax annual median wage (USD)",
                    "totalResults": len(results),
                    "data": results,
                }
                with open(output_file, "w") as f:
                    json.dump(output, f, indent=2, ensure_ascii=False)
                print(f"  ✓ Extracted {len(results)} data points → {output_file}")
                return True
            else:
                print("  ✗ No matching data found")
                return False


# ═══════════════════════════════════════════════════════════════════
#  Phase 2: Eurostat SES (European cities)
# ═══════════════════════════════════════════════════════════════════

def phase2_eurostat():
    """
    Fetch Eurostat Structure of Earnings Survey data.
    Dataset: earn_ses18_28 (or latest SES release)
    ISCO-08 2-digit occupation × country × annual/monthly wages
    """
    print("\n" + "=" * 60)
    print("PHASE 2: Eurostat Structure of Earnings Survey")
    print("=" * 60)

    output_file = RAW_DIR / "eurostat-ses-earnings.json"

    # Eurostat JSON API
    # Dataset: earn_ses18_28 = SES: Structure of Earnings Survey
    # But this may be outdated. Try the newer SDMX format.
    
    # Alternative: Use the bulk download TSV format
    # earn_ses_ann2b = Annual earnings (in EUR) by ISCO-08 occupation, broken by country
    
    datasets_to_try = [
        ("earn_ses_ann2b", "Annual earnings by ISCO-08 2-digit"),
        ("earn_ses18_28", "SES 2018 annual earnings by occupation"),
        ("earn_ses_pub2s", "SES public data by occupation"),
    ]
    
    for dataset_id, desc in datasets_to_try:
        print(f"\n  Trying Eurostat dataset: {dataset_id} ({desc})...")
        
        # Eurostat JSON API format
        url = f"https://ec.europa.eu/eurostat/api/dissemination/statistics/1.0/data/{dataset_id}?format=JSON&lang=EN"
        
        raw = fetch(url, desc, timeout=60)
        if raw:
            try:
                data = json.loads(raw.decode("utf-8"))
                size = len(raw)
                
                # Check if it has useful structure
                dims = data.get("dimension", {})
                values = data.get("value", {})
                
                print(f"  ✓ Got {size:,} bytes")
                print(f"    Dimensions: {list(dims.keys())}")
                print(f"    Values: {len(values)} data points")
                
                # Save raw
                with open(output_file, "w") as f:
                    json.dump(data, f)
                print(f"    Saved → {output_file}")
                
                # Show available dimension values
                for dim_name, dim_data in dims.items():
                    cats = dim_data.get("category", {}).get("label", {})
                    if len(cats) <= 30:
                        print(f"    {dim_name}: {list(cats.values())[:10]}...")
                    else:
                        print(f"    {dim_name}: {len(cats)} categories")
                
                return True
            except json.JSONDecodeError:
                print(f"  ✗ Response is not JSON")
                continue
        else:
            print(f"  ✗ Failed to fetch {dataset_id}")
    
    # Fallback: Try the Eurostat bulk download (TSV/gzip)
    print("\n  Trying Eurostat bulk TSV download...")
    bulk_url = "https://ec.europa.eu/eurostat/api/dissemination/sdmx/2.1/data/earn_ses_ann2b/?format=TSV&compressed=true"
    raw = fetch(bulk_url, "Eurostat TSV", timeout=60)
    if raw:
        import gzip
        try:
            decompressed = gzip.decompress(raw).decode("utf-8")
            tsv_file = RAW_DIR / "eurostat-ses-earnings.tsv"
            with open(tsv_file, "w") as f:
                f.write(decompressed)
            print(f"  ✓ Got Eurostat TSV ({len(decompressed):,} bytes) → {tsv_file}")
            return True
        except Exception as e:
            # Maybe not gzipped
            tsv_file = RAW_DIR / "eurostat-ses-earnings.tsv"
            with open(tsv_file, "wb") as f:
                f.write(raw)
            print(f"  ✓ Got Eurostat data ({len(raw):,} bytes) → {tsv_file}")
            return True
    
    print("  ✗ Eurostat data collection failed")
    return False


# ═══════════════════════════════════════════════════════════════════
#  Phase 3: World Bank — GNI per capita + PPP conversion factors
# ═══════════════════════════════════════════════════════════════════

def phase3_worldbank():
    """
    Fetch World Bank GNI per capita (Atlas method) and PPP conversion factors.
    Used for normalizing salary data across countries.
    """
    print("\n" + "=" * 60)
    print("PHASE 3: World Bank — GNI per capita & PPP")
    print("=" * 60)

    indicators = {
        "NY.GNP.PCAP.CD": "GNI per capita (Atlas, current USD)",
        "NY.GNP.PCAP.PP.CD": "GNI per capita (PPP, current intl $)",
        "PA.NUS.PPPC.RF": "PPP conversion factor (GDP)",
    }
    
    for indicator, desc in indicators.items():
        print(f"\n  Fetching {desc}...")
        url = f"https://api.worldbank.org/v2/country/all/indicator/{indicator}?date=2022:2024&format=json&per_page=500"
        
        raw = fetch(url, desc, timeout=30)
        if raw:
            try:
                data = json.loads(raw.decode("utf-8"))
                
                # World Bank returns [metadata, data]
                if isinstance(data, list) and len(data) >= 2:
                    records = data[1] if data[1] else []
                    print(f"  ✓ {len(records)} records")
                    
                    fname = f"worldbank-{indicator.replace('.', '-').lower()}.json"
                    with open(RAW_DIR / fname, "w") as f:
                        json.dump(records, f, indent=2)
                    print(f"    Saved → {fname}")
                    
                    # Show sample
                    if records:
                        us = next((r for r in records if r.get("country", {}).get("id") == "US" and r.get("value")), None)
                        if us:
                            print(f"    US sample: {us['date']} = {us['value']}")
                else:
                    print(f"  ✗ Unexpected format: {type(data)}")
            except json.JSONDecodeError:
                print(f"  ✗ Not JSON")
        else:
            print(f"  ✗ Failed")


# ═══════════════════════════════════════════════════════════════════
#  Phase 4: Process existing ILO data
# ═══════════════════════════════════════════════════════════════════

def phase4_ilo():
    """
    Structure the already-collected ILO data into a usable format.
    ILO gives us:
      - National average monthly earnings by ISCO-08 major group (1-digit)
      - This provides the inter-occupation ratio per country
    """
    print("\n" + "=" * 60)
    print("PHASE 4: Process ILO occupation data")
    print("=" * 60)

    ilo_file = RAW_DIR / "ilo-ilo-wages-by-occupation-sdg.json"
    if not ilo_file.exists():
        print(f"  ✗ ILO data file not found: {ilo_file}")
        return False
    
    with open(ilo_file) as f:
        raw_data = json.load(f)
    
    print(f"  Loaded {len(raw_data):,} ILO records")
    
    # Extract: country × ISCO major group × latest year → monthly earnings (local currency)
    # ISCO-08 major groups that map to our professions:
    ISCO_GROUP_MAP = {
        "1. Managers": ["市场经理", "销售经理", "人力资源经理", "产品经理"],
        "2. Professionals": ["软件工程师", "医生/医学博士", "财务分析师", "数据科学家",
                             "律师", "建筑师", "机械工程师", "药剂师", "会计师",
                             "UI/UX设计师", "大学教授", "牙医", "记者", "护士"],
        "3. Technicians and associate professionals": ["平面设计师"],
        "5. Service and sales workers": ["厨师", "政府/NGO行政"],
        "7. Craft and related trades workers": ["电工"],
        "8. Plant and machine operators, and assemblers": ["公交司机"],
        "9. Elementary occupations": [],  # reference baseline
    }
    
    # Filter for ISCO-08 data, Total sex, latest year per country
    country_data = {}
    for item in raw_data:
        classif = item.get("classif1.label", "")
        if "ISCO-08" not in classif:
            continue
        
        country = item["ref_area.label"]
        year = int(item["time"]) if str(item.get("time", "")).isdigit() else 0
        value = item.get("obs_value")
        sex = item.get("sex.label", "")
        
        if sex != "Total" or value is None or year < 2018:
            continue
        
        # Extract ISCO group name
        isco_group = classif.replace("Occupation (ISCO-08): ", "")
        
        if country not in country_data:
            country_data[country] = {}
        
        key = (isco_group, year)
        if key not in country_data[country]:
            country_data[country][key] = value
        # Keep latest year
    
    # Restructure: per country, get the latest year's ISCO ratios
    output = {}
    for country, entries in country_data.items():
        # Find latest year with 'Total'
        totals = {yr: val for (grp, yr), val in entries.items() if grp == "Total"}
        if not totals:
            continue
        latest_year = max(totals.keys())
        total_val = totals[latest_year]
        
        if total_val is None or total_val <= 0:
            continue
        
        groups = {}
        for (grp, yr), val in entries.items():
            if yr == latest_year and val and val > 0:
                groups[grp] = {
                    "monthlyEarnings_local": val,
                    "ratioToTotal": round(val / total_val, 3),
                }
        
        if len(groups) > 2:  # at least Total + 1 group
            output[country] = {
                "year": latest_year,
                "totalMonthly_local": total_val,
                "groups": groups,
            }
    
    output_file = RAW_DIR / "ilo-occupation-ratios.json"
    with open(output_file, "w") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)
    
    print(f"  ✓ Extracted occupation ratios for {len(output)} countries → {output_file}")
    
    # Show sample
    for sample_country in ["United States of America", "United Kingdom of Great Britain and Northern Ireland", "Japan", "Germany"]:
        if sample_country in output:
            d = output[sample_country]
            print(f"\n  {sample_country} ({d['year']}):")
            for grp, info in sorted(d["groups"].items()):
                print(f"    {grp:55s} ratio={info['ratioToTotal']:.2f}")
    
    return True


# ═══════════════════════════════════════════════════════════════════
#  Main
# ═══════════════════════════════════════════════════════════════════

def main():
    print("╔══════════════════════════════════════════════════════════╗")
    print("║  WhichCity Salary Data Collection — Phase 1 (Open APIs) ║")
    print("╚══════════════════════════════════════════════════════════╝")
    print(f"\nOutput directory: {RAW_DIR}")
    print(f"Flags: {'--skip-bls' if SKIP_BLS else ''} {'--bls-only' if BLS_ONLY else ''}")
    
    results = {}
    
    if not SKIP_BLS:
        # Check if zip was manually downloaded
        if (RAW_DIR / "oesm24ma.zip").exists() and not (RAW_DIR / "bls-oews-extracted.json").exists():
            results["BLS (zip parse)"] = phase1b_parse_bls_zip()
        else:
            results["BLS"] = phase1_bls()
            # If text format failed but zip exists, try parsing zip
            if not results.get("BLS") and (RAW_DIR / "oesm24ma.zip").exists():
                results["BLS (zip parse)"] = phase1b_parse_bls_zip()
    
    if not BLS_ONLY:
        results["Eurostat"] = phase2_eurostat()
        results["World Bank"] = phase3_worldbank()
        results["ILO restructure"] = phase4_ilo()
    
    # Summary
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    for name, ok in results.items():
        status = "✓" if ok else "✗"
        print(f"  {status} {name}")
    
    print(f"\nRaw files in {RAW_DIR}:")
    for f in sorted(RAW_DIR.iterdir()):
        if f.is_file():
            size = f.stat().st_size
            print(f"  {f.name:50s} {size:>12,} bytes")
    
    print("\n╔══════════════════════════════════════════════════════════╗")
    print("║  Next: Run collect-salary-proxy.mjs for Numbeo data     ║")
    print("║  (requires proxy/IP rotation, see instructions below)   ║")
    print("╚══════════════════════════════════════════════════════════╝")


if __name__ == "__main__":
    main()
